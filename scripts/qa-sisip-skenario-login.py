# -*- coding: utf-8 -*-
"""
Selaraskan sheet QA manual dengan keadaan setelah B-27 (login Identity Platform).

Berkas tujuan: docs/QA-Manual-AgroVision-SprintA-3-3B-20260826.xlsx.
Dua langkah, dua sheet; sheet "Log Bug" TIDAK disentuh sama sekali:

  LANGKAH 1 — sheet "Skenario QA": sisipkan kelompok skenario login L-01..L-07.
    Sumber teks: docs/16-prioritas-qa-20260826.md, bagian "Tier 0 · Login".

  LANGKAH 2 — sheet "Petunjuk": perbaiki kalimat yang sudah usang sejak B-27
    mendarat 26 Agustus 2026 (URL aplikasi, judul blok akun uji, baris
    "Login stub", urutan kerja yang belum menyebut kelompok L) dan tambahkan
    satu baris "Kredensial" di blok CATATAN PENTING.
    Sumber: CLAUDE.md bagian "Autentikasi — dua mode" + docs/12-deploy-gcp.md §9.

Kenapa lewat skrip, bukan diedit tangan: perubahannya bisa diulang, bisa
diperiksa orang lain, dan idempoten — kedua langkah berhenti sendiri kalau
hasilnya sudah ada di berkas. Langkah 2 juga menolak menebak: kalau teks lama
di satu sel tidak sama persis dengan yang diharapkan, skrip berhenti dan
melapor, bukan menimpa.

Jalankan:
    pip3 install --user openpyxl        # sekali saja
    python3 scripts/qa-sisip-skenario-login.py           # menulis
    python3 scripts/qa-sisip-skenario-login.py --cek     # hanya memeriksa

ATURAN YANG TIDAK BOLEH DILANGGAR
- JANGAN pernah menulis kata sandi ke dalam berkas Excel ini. Sheet QA
  beredar lewat lampiran chat. Skenario yang butuh akun menyebut PERANNYA
  ("akun peran creator"), bukan sandinya; sandi ada di password manager tim
  (dikirim Dimas terpisah, lihat docs/12-deploy-gcp.md §9).
- Login stub sudah tidak berlaku di produksi (butuh tiga gerbang sekaligus:
  AUTH_MODE=stub, NODE_ENV != production, dan saklar DB
  app.auth_settings.stub_login_enabled). Skenario di bawah adalah skenario
  Identity Platform — memakai kata sandi sungguhan.

ASUMSI STRUKTUR BERKAS (diverifikasi ulang saat skrip jalan, lihat
`periksa_struktur`); kalau sheet-nya berubah, skrip berhenti, tidak menebak:

  Sheet "Skenario QA", 15 kolom, header di baris 1:
    A(1)  ID                  contoh "A-01"; huruf = kelompok, angka = urutan
    B(2)  Menu (grup)         grup menu di sidebar, mis. "Autentikasi"
    C(3)  Modul / Layar       nama layar, mis. "Login"
    D(4)  Path URL            mis. "/login"; "(berbagai)" bila lintas layar
    E(5)  Role                peran penguji, mis. "Creator", "Semua"
    F(6)  Skenario            satu kalimat: apa yang diuji
    G(7)  Langkah Uji         "1) ... 2) ... 3) ..."
    H(8)  Data Uji            data/akun yang dipakai; "—" bila tidak ada
    I(9)  Hasil yang Diharapkan
    J(10) Prioritas           Critical / High / Medium / Low
    K(11) Status              DIISI QA — dropdown PASS/FAIL/BLOCKED/SKIP
    L(12) Perangkat Diuji     DIISI QA — dropdown perangkat
    M(13) Tanggal             DIISI QA
    N(14) Catatan             DIISI QA (boleh sudah terisi catatan penyusun)
    O(15) Ref Bug             DIISI QA — nomor bug di sheet "Log Bug"

  Dua jenis baris:
    - baris KEPALA KELOMPOK: kolom A = huruf kelompok ("A", "B", ... "H"),
      kolom B = judul kelompok, kolom C..O kosong; latar abu tua, teks putih.
    - baris DATA: kolom A = "X-01"; latar polos, rata atas, bungkus teks.
  Kolom K dan L punya data validation (dropdown) yang sqref-nya harus
  diperpanjang sampai baris terakhir yang baru, begitu juga auto filter.
  Tidak ada tinggi baris kustom, merged cell, gambar, atau komentar — jadi
  menyimpan ulang dengan openpyxl tidak menghilangkan apa pun.

Kelompok baru memakai huruf "L" (Login) dan DITAMBAHKAN DI BAWAH kelompok H,
bukan disisipkan di atas: nomor baris skenario lama tidak boleh bergeser
(QA sudah memakai nomor baris saat berdiskusi). Urutan pengerjaannya —
kelompok L dikerjakan PALING AWAL — ditulis di judul kelompoknya, karena
sheet "Petunjuk" tidak boleh diubah dari skrip ini.
"""

import os
import sys
from copy import copy

try:
    import openpyxl
except ImportError:  # pragma: no cover - pesan lebih berguna daripada traceback
    sys.exit("openpyxl belum terpasang. Jalankan: pip3 install --user openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "QA-Manual-AgroVision-SprintA-3-3B-20260826.xlsx")
SHEET = "Skenario QA"

# Header yang diharapkan, kolom 1..15. Dipakai sebagai pagar: kalau tidak
# sama persis, struktur sheet sudah berubah dan skrip ini tidak lagi valid.
HEADER = [
    "ID", "Menu (grup)", "Modul / Layar", "Path URL", "Role", "Skenario",
    "Langkah Uji", "Data Uji", "Hasil yang Diharapkan", "Prioritas",
    "Status", "Perangkat Diuji", "Tanggal", "Catatan", "Ref Bug",
]

KEPALA_KELOMPOK = (
    "L",
    "B-27 · LOGIN IDENTITY PLATFORM & SESI (sudah live 26 Agustus — "
    "TIER 0, KERJAKAN PALING AWAL SEBELUM KELOMPOK A)",
)

# Tujuh baris skenario. Urutan nilai = urutan kolom A..J; kolom K..M dan O
# sengaja dibiarkan kosong (diisi QA), kolom N (Catatan) diisi hanya bila ada
# hal yang perlu diketahui walau statusnya PASS — mengikuti kebiasaan sheet.
# Teksnya diambil dari docs/16-prioritas-qa-20260826.md; kolom "Data Uji" dan
# "Catatan" tidak ada di dokumen itu dan ditulis di sini.
BARIS = [
    {
        "ID": "L-01",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Login",
        "Path URL": "/login",
        "Role": "Semua (4 akun)",
        "Skenario": "Login dengan kata sandi yang benar",
        "Langkah Uji": (
            "1) Buka /login 2) Isi email + kata sandi 3) Tekan Masuk "
            "4) Perhatikan nama & label peran di kanan atas 5) Logout, lalu "
            "ulangi untuk keempat akun demo"
        ),
        "Data Uji": (
            "4 akun demo (admin@, approver@, creator@, direktur@demo.invalid). "
            "Kata sandinya dikirim Dimas terpisah dan disimpan di password "
            "manager tim — JANGAN ditulis di sheet ini"
        ),
        "Hasil yang Diharapkan": (
            "Keempatnya mendarat di Dashboard sesuai perannya, dengan label "
            "peran yang benar (Super Admin / Approver / Petugas Lapangan / "
            "Viewer). Satu akun saja gagal masuk → FAIL Critical, dan 72 "
            "skenario lain ikut terblokir"
        ),
        "Prioritas": "Critical",
        "Catatan": (
            "Kerjakan paling awal (Tier 0, docs/16-prioritas-qa-20260826.md). "
            "Instruksi di sheet 'Petunjuk' — 'login tanpa password, cukup ketik "
            "email' — sudah TIDAK berlaku sejak B-27 mendarat 26 Agustus. "
            "Alamat @demo.invalid tidak bisa menerima email, jadi tidak ada "
            "alur reset kata sandi: sandi hilang = akun dibuat ulang oleh Dimas."
        ),
    },
    {
        "ID": "L-02",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Login",
        "Path URL": "/login",
        "Role": "Creator",
        "Skenario": "Kata sandi salah ditolak",
        "Langkah Uji": (
            "1) Buka /login 2) Isi email creator@ yang BENAR + kata sandi asal "
            "3) Tekan Masuk 4) Salin kalimat pesan galatnya untuk dibandingkan "
            "di L-03"
        ),
        "Data Uji": "Akun peran creator, kata sandi diketik asal",
        "Hasil yang Diharapkan": (
            "Ditolak dengan pesan persis 'Email atau kata sandi salah.' dan "
            "tetap di halaman login"
        ),
        "Prioritas": "Critical",
        "Catatan": "",
    },
    {
        "ID": "L-03",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Login",
        "Path URL": "/login",
        "Role": "—",
        "Skenario": "Email tak terdaftar ditolak dengan pesan YANG SAMA PERSIS",
        "Langkah Uji": (
            "1) Buka /login 2) Isi bukansiapa@demo.invalid + kata sandi asal "
            "3) Tekan Masuk 4) Bandingkan kalimat pesannya kata per kata "
            "dengan yang dicatat di L-02"
        ),
        "Data Uji": "bukansiapa@demo.invalid — email yang memang tidak terdaftar",
        "Hasil yang Diharapkan": (
            "Pesannya IDENTIK dengan L-02 ('Email atau kata sandi salah.'). "
            "Kalau berbeda sedikit pun → FAIL: layar login jadi bisa dipakai "
            "menebak email mana yang terdaftar (enumerasi akun)"
        ),
        "Prioritas": "Critical",
        "Catatan": "",
    },
    {
        "ID": "L-04",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Login",
        "Path URL": "/login",
        "Role": "—",
        "Skenario": "Akun terverifikasi tapi belum ditautkan ke pengguna AgroVision",
        "Langkah Uji": (
            "1) Minta Dimas membuat satu akun Identity Platform sekali pakai "
            "TANPA menautkan external_id 2) Login dengan akun itu"
        ),
        "Data Uji": (
            "Akun Identity Platform sekali pakai tanpa external_id — minta ke "
            "Dimas; kata sandinya lewat password manager tim, bukan lewat sheet"
        ),
        "Hasil yang Diharapkan": (
            "Ditolak dengan pesan yang menyebut 'belum terhubung ke pengguna "
            "AgroVision' — BUKAN 'Email atau kata sandi salah.' Kredensialnya "
            "memang benar; yang belum ada penautannya"
        ),
        "Prioritas": "High",
        "Catatan": (
            "Penautan external_id sengaja jadi tindakan admin lewat koneksi "
            "superuser (docs/12-deploy-gcp.md §9 langkah 3), bukan sesuatu yang "
            "dikerjakan sendiri oleh login pertama."
        ),
    },
    {
        "ID": "L-05",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Login",
        "Path URL": "/login",
        "Role": "Semua",
        "Skenario": "Produksi tidak lagi mengumumkan mode pengembangan",
        "Langkah Uji": (
            "1) Buka /login di URL produksi 2) Perhatikan seluruh halaman dari "
            "atas sampai bawah, tanpa login"
        ),
        "Data Uji": "— (cukup halaman /login)",
        "Hasil yang Diharapkan": (
            "TIDAK ADA kotak kuning 'Mode pengembangan', dan ADA kolom Kata "
            "sandi. Kotak kuning itu masih muncul berarti login stub menyala di "
            "produksi → FAIL Critical"
        ),
        "Prioritas": "Critical",
        "Catatan": "",
    },
    {
        "ID": "L-06",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Logout & sesi",
        "Path URL": "/dashboard",
        "Role": "Creator",
        "Skenario": "Logout benar-benar memutus sesi",
        "Langkah Uji": (
            "1) Login dengan akun peran creator 2) Logout 3) Tempel URL "
            "/dashboard langsung di bilah alamat"
        ),
        "Data Uji": "Akun peran creator",
        "Hasil yang Diharapkan": (
            "Dialihkan ke /login, bukan menampilkan dashboard"
        ),
        "Prioritas": "High",
        "Catatan": (
            "Aplikasi ini PWA: kalau dashboard sempat tampil, tutup TOTAL "
            "peramban lalu ulangi — bisa jadi service worker menyajikan versi "
            "lama, bukan sesi yang masih hidup."
        ),
    },
    {
        "ID": "L-07",
        "Menu (grup)": "Autentikasi",
        "Modul / Layar": "Penonaktifan",
        "Path URL": "/pengguna",
        "Role": "Super Admin > Creator",
        "Skenario": "Menonaktifkan pengguna langsung berlaku pada sesi yang sedang jalan",
        "Langkah Uji": (
            "1) Login sebagai creator di satu peramban, biarkan terbuka "
            "2) Di peramban lain, super admin buka /pengguna dan nonaktifkan "
            "creator itu 3) Kembali ke peramban creator, muat ulang halaman"
        ),
        "Data Uji": (
            "Akun peran creator + akun peran super admin, di dua peramban "
            "berbeda (bukan dua tab)"
        ),
        "Hasil yang Diharapkan": (
            "Creator langsung kehilangan akses tanpa perlu logout — sesi "
            "diperiksa ulang ke database tiap request. Masih bisa menjelajah "
            "sampai logout → FAIL"
        ),
        "Prioritas": "High",
        "Catatan": (
            "Pelengkap E-04: E-04 menguji dari sisi login BARU, L-07 dari sisi "
            "sesi yang SUDAH berjalan. Aktifkan kembali akun creator setelah "
            "uji selesai, karena skenario kelompok lain memakainya."
        ),
    },
]


def periksa_struktur(ws):
    """Berhenti kalau sheet tidak lagi seperti yang diasumsikan di atas."""
    aktual = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if aktual != HEADER:
        sys.exit(
            "Struktur sheet berubah — header baris 1 tidak sama dengan asumsi "
            "skrip.\n  diharapkan: %s\n  ditemukan : %s" % (HEADER, aktual)
        )


def id_yang_ada(ws):
    """Semua nilai kolom ID (kolom A) sebagai daftar string."""
    return [
        str(ws.cell(row=r, column=1).value or "").strip()
        for r in range(2, ws.max_row + 1)
    ]


def salin_gaya(sumber, tujuan):
    """Tiru gaya satu sel: font, latar, garis, perataan, format angka.

    Dipakai supaya baris baru tidak perlu mendefinisikan warna/garis sendiri —
    kalau sheet-nya nanti diberi tema lain, baris L ikut berubah dengan
    sendirinya karena gayanya dicontek dari baris yang sudah ada.
    """
    tujuan.font = copy(sumber.font)
    tujuan.fill = copy(sumber.fill)
    tujuan.border = copy(sumber.border)
    tujuan.alignment = copy(sumber.alignment)
    tujuan.number_format = sumber.number_format


def perpanjang_rentang(ws, baris_terakhir):
    """Perpanjang dropdown (kolom Status & Perangkat Diuji) dan auto filter.

    openpyxl tidak melakukannya sendiri: tanpa ini, baris baru tampil tanpa
    dropdown dan tidak ikut tersaring saat QA memfilter kolom.
    """
    for dv in ws.data_validations.dataValidation:
        rentang = str(dv.sqref)
        for kolom in ("K", "L"):
            if rentang.startswith(kolom + "2:"):
                dv.sqref = "%s2:%s%d" % (kolom, kolom, baris_terakhir)
    if ws.auto_filter.ref:
        ws.auto_filter.ref = "A1:%s%d" % (
            openpyxl.utils.get_column_letter(ws.max_column), baris_terakhir
        )


def cetak_baris(ws, nomor_baris):
    """Cetak isi baris ke layar — dipakai untuk verifikasi setelah menulis."""
    for r in nomor_baris:
        print("-" * 78)
        for c in range(1, ws.max_column + 1):
            nilai = ws.cell(row=r, column=c).value
            if nilai in (None, ""):
                continue
            print("  R%-3d %-22s | %s" % (r, HEADER[c - 1], nilai))


# --------------------------------------------------------------------------
# LANGKAH 2 — sheet "Petunjuk"
#
# Strukturnya prosa, bukan tabel: kolom A = label (atau judul blok, bercetak
# tebal hijau FF1B5E20 dan dibiarkan meluber ke kolom B yang kosong), kolom B =
# isinya (lebar 96, bungkus teks). Tanpa merged cell, tanpa dropdown, tanpa
# tinggi baris kustom — jadi mengganti .value satu sel tidak mengubah gayanya
# sama sekali, dan tidak ada baris yang perlu digeser.
#
# Setiap perubahan ditulis sebagai (koordinat, teks lama PERSIS, teks baru).
# Teks lama dipakai sebagai kunci pengaman: kalau isinya sudah teks baru →
# lewati (idempoten); kalau bukan keduanya → berhenti, karena berarti ada orang
# yang sudah menyuntingnya dan skrip ini tidak boleh menimpa suntingan itu.
# --------------------------------------------------------------------------

PERUBAHAN_PETUNJUK = [
    (
        # URL yang dipublikasikan. Bentuk lama (berbasis nomor project) boleh
        # jadi masih ikut menjawab — Cloud Run kerap punya dua bentuk URL —
        # tapi yang tercantum di sheet cukup satu, yang ini.
        "B7",
        "https://agrovision-393569486275.asia-southeast2.run.app",
        "https://agrovision-pjy4ku3jjq-et.a.run.app",
    ),
    (
        # Judul blok akun uji. Daftar emailnya di bawah (A30:B34) DIBIARKAN —
        # email akun demo bukan rahasia; yang salah hanya janji "cukup ketik
        # email". Kata sandinya tidak ditulis di sini maupun di mana pun dalam
        # berkas ini; penunjuknya ada di baris "Kredensial" (lihat
        # BARIS_CATATAN_BARU).
        "A29",
        "AKUN UJI (login tanpa password, cukup ketik email)",
        "AKUN UJI (login Identity Platform — wajib kata sandi sejak B-27)",
    ),
    (
        # Urutan kerja: kelompok L (login) baru ditambahkan dan letaknya di
        # BAWAH kelompok H, jadi "A sampai H" akan membuat penguji melewatinya.
        "B42",
        "Buka sheet 'Skenario QA'. Kerjakan berurutan per kelompok (A sampai H).",
        "Buka sheet 'Skenario QA'. Kerjakan kelompok L (login, ada di paling "
        "bawah) DULU, baru A sampai H berurutan.",
    ),
    (
        # Label baris batasan: "Login stub" bukan lagi keadaan produksi.
        "A51",
        "Login stub",
        "Autentikasi",
    ),
    (
        "B51",
        "Autentikasi belum memverifikasi kredensial — cukup email terdaftar. "
        "Jangan masukkan data sungguhan.",
        "Produksi MEMVERIFIKASI kredensial (B-27): peramban menukar kata sandi "
        "langsung ke Identity Platform, server hanya menerima ID token lalu "
        "memeriksa tanda tangan RS256 + iss/aud/exp/iat sebelum sesi dibuat — "
        "kata sandi tidak pernah melewati server aplikasi. Login stub (email "
        "tanpa kata sandi) masih ada TAPI hanya untuk pengembangan lokal & "
        "at:verify, dan butuh tiga gerbang sekaligus: AUTH_MODE=stub, "
        "NODE_ENV != production, dan saklar database "
        "app.auth_settings.stub_login_enabled. Rujukan: CLAUDE.md "
        "\"Autentikasi — dua mode\" dan docs/12-deploy-gcp.md §9.",
    ),
]

# Satu baris tambahan di blok CATATAN PENTING (blok terakhir di sheet, jadi
# cukup ditambahkan di bawahnya — tidak ada baris yang bergeser). Gayanya
# dicontek dari baris "Data demo" (A49/B49).
BARIS_CATATAN_BARU = (
    "Kredensial",
    "Kata sandi akun demo dikirim Dimas terpisah dan disimpan di password "
    "manager tim — jangan pernah ditulis di sheet ini. Alamat @demo.invalid "
    "tidak bisa menerima email, jadi tidak ada alur reset kata sandi: sandi "
    "hilang berarti akunnya dibuat ulang oleh Dimas.",
)
BARIS_CONTEKAN_CATATAN = 49  # baris "Data demo" — sumber gaya


def perbarui_petunjuk(wb, hanya_cek=False):
    """Perbaiki kalimat usang di sheet "Petunjuk".

    Mengembalikan daftar (koordinat, teks lama, teks baru) yang benar-benar
    berubah. Daftar kosong = semuanya sudah mutakhir.
    """
    ws = wb["Petunjuk"]
    berubah = []

    for koordinat, lama, baru in PERUBAHAN_PETUNJUK:
        sel = ws[koordinat]
        nilai = sel.value
        if nilai == baru:
            continue  # sudah diperbarui pada jalan sebelumnya
        if nilai != lama:
            sys.exit(
                "Sel Petunjuk!%s tidak berisi teks lama yang diharapkan — "
                "berhenti, tidak menimpa.\n  diharapkan: %r\n  ditemukan : %r"
                % (koordinat, lama, nilai)
            )
        if not hanya_cek:
            sel.value = baru
        berubah.append((koordinat, lama, baru))

    # Baris "Kredensial" — idempoten lewat pencarian labelnya di kolom A.
    label_baru, isi_baru = BARIS_CATATAN_BARU
    sudah_ada = any(
        ws.cell(row=r, column=1).value == label_baru
        for r in range(1, ws.max_row + 1)
    )
    if not sudah_ada:
        r = ws.max_row + 1
        if not hanya_cek:
            for kolom, nilai in ((1, label_baru), (2, isi_baru)):
                sel = ws.cell(row=r, column=kolom)
                salin_gaya(ws.cell(row=BARIS_CONTEKAN_CATATAN, column=kolom), sel)
                sel.value = nilai
        berubah.append(("A%d/B%d" % (r, r), "(baris baru)", "%s | %s" % (label_baru, isi_baru)))

    return berubah


def sisip_skenario(wb, hanya_cek=False):
    """Sisipkan kepala kelompok L + tujuh baris skenario di sheet "Skenario QA".

    Mengembalikan daftar nomor baris yang ditambahkan; daftar kosong = kelompok
    L sudah ada dan tidak ada yang ditulis.
    """
    ws = wb[SHEET]
    periksa_struktur(ws)

    sudah = [i for i in id_yang_ada(ws) if i == "L" or i.startswith("L-")]
    if sudah:
        print("[1/2] Skenario QA: kelompok L sudah ada (%s) — dilewati."
              % ", ".join(sudah))
        return []
    if hanya_cek:
        print("[1/2] Skenario QA: kelompok L BELUM ada, %d baris siap disisipkan."
              % len(BARIS))
        return []

    # Baris contekan gaya: baris 2 = kepala kelompok "A", baris 3 = data "A-01".
    gaya_kepala = 2
    gaya_data = 3

    baris_kepala = ws.max_row + 1
    for c in range(1, ws.max_column + 1):
        sel = ws.cell(row=baris_kepala, column=c)
        salin_gaya(ws.cell(row=gaya_kepala, column=c), sel)
    ws.cell(row=baris_kepala, column=1).value = KEPALA_KELOMPOK[0]
    ws.cell(row=baris_kepala, column=2).value = KEPALA_KELOMPOK[1]

    baris_baru = [baris_kepala]
    for data in BARIS:
        r = ws.max_row + 1
        for c in range(1, ws.max_column + 1):
            sel = ws.cell(row=r, column=c)
            salin_gaya(ws.cell(row=gaya_data, column=c), sel)
            nilai = data.get(HEADER[c - 1], "")
            if nilai:
                sel.value = nilai
        baris_baru.append(r)

    perpanjang_rentang(ws, ws.max_row)
    print("[1/2] Skenario QA: 1 kepala kelompok + %d skenario (baris %d-%d)."
          % (len(BARIS), baris_baru[0], baris_baru[-1]))
    return baris_baru


def main():
    hanya_cek = "--cek" in sys.argv

    wb = openpyxl.load_workbook(XLSX)
    baris_baru = sisip_skenario(wb, hanya_cek)
    petunjuk_berubah = perbarui_petunjuk(wb, hanya_cek)

    if petunjuk_berubah:
        print("[2/2] Petunjuk: %d sel diperbarui." % len(petunjuk_berubah))
        for koordinat, lama, baru in petunjuk_berubah:
            print("      %s\n        lama: %s\n        baru: %s" % (koordinat, lama, baru))
    else:
        print("[2/2] Petunjuk: sudah mutakhir — dilewati.")

    if hanya_cek:
        print("(--cek: tidak ada yang ditulis ke berkas)")
        return 0
    if not baris_baru and not petunjuk_berubah:
        print("Tidak ada perubahan. Berkas tidak ditulis ulang.")
        return 0

    wb.save(XLSX)
    print("Tersimpan: %s" % os.path.basename(XLSX))

    # Verifikasi: baca ULANG dari disk, bukan dari objek yang barusan ditulis.
    ulang = openpyxl.load_workbook(XLSX)
    if baris_baru:
        cetak_baris(ulang[SHEET], baris_baru)
    if petunjuk_berubah:
        print("-" * 78)
        print("  Sheet 'Petunjuk' setelah perubahan:")
        wsp = ulang["Petunjuk"]
        for koordinat, _, _ in petunjuk_berubah:
            for satu in koordinat.split("/"):
                print("  %-5s %r" % (satu, wsp[satu].value))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
